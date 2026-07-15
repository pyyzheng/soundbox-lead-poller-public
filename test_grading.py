#!/usr/bin/env python3
"""
test_grading.py — 线索分级质量验证工具（模块化版）

功能：
  1. 从 Excel 或 JSONL 读取 Google 渠道历史询盘
  2. 逐条调用 grade_lead()（代码规则 + LLM 语义槽位）
  3. 输出详细报告（每条的 Level + 槽位 + 信号 + 原文摘要）
  4. 如果有人工标签，计算准确率和召回率

Usage:
  ZHIPU_API_KEY=xxx python test_grading.py --source 询盘分配表_已清洗.xlsx
  ZHIPU_API_KEY=xxx python test_grading.py --source 询盘分配表_已清洗.xlsx --limit 10
  ZHIPU_API_KEY=xxx python test_grading.py --source 询盘分配表_已清洗.xlsx --channel-only google
  ZHIPU_API_KEY=xxx python test_grading.py --source 询盘分配表_已清洗.xlsx --labels manual_labels.json

  # 从 stdin 读取单条测试
  echo "I need 10 office pods..." | ZHIPU_API_KEY=xxx python test_grading.py --stdin
"""

import os
import sys
import json
import time
import argparse
import logging
from pathlib import Path
from datetime import datetime

# 添加 lib 目录到 path
sys.path.insert(0, str(Path(__file__).parent / "lib"))

from lead_grader import grade_lead

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("test-grading")


# ── 数据读取 ─────────────────────────────────────────────────────────────────

def load_from_excel(filepath: str, channel_only: str | None = None) -> list[dict]:
    """从 Excel 文件读取询盘数据。"""
    try:
        import openpyxl
    except ImportError:
        log.error("需要 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # 读取表头
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]

    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = val

        # 渠道过滤
        ch = str(record.get("渠道", "") or "")
        if channel_only == "google":
            if "谷歌" not in ch and "google" not in ch.lower():
                continue

        # 需要有询盘内容
        content = record.get("询盘内容", "")
        if not content or not str(content).strip():
            continue

        leads.append(record)

    wb.close()
    log.info(f"从 Excel 加载 {len(leads)} 条询盘（渠道过滤: {channel_only}）")
    return leads


def load_from_jsonl(filepath: str) -> list[dict]:
    """从 JSONL 文件读取询盘数据。"""
    leads = []
    with open(filepath, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            leads.append(json.loads(line))
    log.info(f"从 JSONL 加载 {len(leads)} 条询盘")
    return leads


def load_manual_labels(filepath: str) -> dict[str, str]:
    """加载人工标签文件。"""
    if filepath.endswith(".json"):
        with open(filepath, encoding="utf-8") as f:
            return json.load(f)
    elif filepath.endswith(".csv"):
        labels = {}
        with open(filepath, encoding="utf-8") as f:
            for line in f:
                parts = line.strip().split(",", 1)
                if len(parts) == 2:
                    labels[parts[0].strip()] = parts[1].strip()
        return labels
    else:
        log.error(f"不支持的标签文件格式: {filepath}")
        return {}


# ── 分级执行 ─────────────────────────────────────────────────────────────────

def grade_leads(leads: list[dict], delay: float = 0.5) -> list[dict]:
    """对每条询盘执行分级（代码规则 + LLM 语义槽位）。"""
    results = []
    total = len(leads)

    for i, lead in enumerate(leads):
        content = str(lead.get("询盘内容", ""))
        contact = str(lead.get("联系人", "") or "").strip()
        email = str(lead.get("Email_norm", lead.get("email", "")) or "").strip()
        channel = str(lead.get("渠道", "") or "")

        log.info(f"[{i+1}/{total}] 分级中: {contact or 'N/A'} ({channel})")

        # 截断超长内容
        if len(content) > 3000:
            content = content[:3000] + "\n...(truncated)"

        t0 = time.time()
        grading = grade_lead(content, email=email)
        elapsed = time.time() - t0

        results.append({
            "index": i + 1,
            "contact": contact,
            "email": email,
            "channel": channel,
            "country": str(lead.get("国家", "") or ""),
            "product": str(lead.get("文本 15", "") or ""),
            "content_preview": content[:200].replace("\n", " "),
            "grading": grading,
            "elapsed": elapsed,
        })

        if delay > 0 and i < total - 1:
            time.sleep(delay)

    return results


# ── 报告生成 ─────────────────────────────────────────────────────────────────

def generate_report(results: list[dict], labels: dict | None = None) -> str:
    """生成分级报告。"""
    lines = []
    lines.append("=" * 80)
    lines.append(f"线索分级验证报告 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"样本数: {len(results)}")
    if results:
        avg_time = sum(r["elapsed"] for r in results) / len(results)
        lines.append(f"平均耗时: {avg_time:.1f}s/条")
    lines.append("=" * 80)

    # ── 分级分布 ──
    level_counts = {}
    for r in results:
        g = r.get("grading")
        level = g.get("level", "FAILED") if g else "FAILED"
        level_counts[level] = level_counts.get(level, 0) + 1

    lines.append("\n## 分级分布")
    lines.append("-" * 40)
    for level in ["Level 1", "Level 2", "Level 3", "Level 4", "FAILED"]:
        count = level_counts.get(level, 0)
        pct = count / len(results) * 100 if results else 0
        bar = "#" * count
        lines.append(f"  {level:10s}: {count:3d} ({pct:5.1f}%) {bar}")

    # ── 逐条详情 ──
    lines.append(f"\n## 逐条详情 ({len(results)} 条)")
    lines.append("-" * 80)

    for r in results:
        g = r.get("grading")
        if g:
            level = g.get("level", "?")
            l2_tag = g.get("l2_tag", "N/A")
            slots = g.get("slots", {})
            sig = g.get("signals", {})
            conclusion = g.get("conclusion", "")
            basis1 = g.get("basis_1_intent_hit", "")
            basis2 = g.get("basis_2_structure_hit", "")
            price_only = g.get("price_only_flag", "")
            active_sig = ", ".join(k for k, v in sig.items() if v) or "none"
        else:
            level = "FAILED"
            l2_tag = slots = conclusion = basis1 = basis2 = price_only = active_sig = ""

        # 人工标签对比
        manual = ""
        if labels:
            email_key = (r.get("email") or "").lower().strip()
            contact_key = (r.get("contact") or "").strip()
            ml = labels.get(email_key) or labels.get(contact_key, "")
            if ml:
                manual = f" | 人工={ml}"
                manual += " ✓" if ml == level else " ✗ MISMATCH"

        lines.append(f"\n[{r['index']}] {level}{manual} ({r['elapsed']:.1f}s)")
        lines.append(f"    联系人: {r.get('contact', 'N/A')}")
        lines.append(f"    邮箱: {r.get('email', 'N/A')}")
        lines.append(f"    渠道: {r.get('channel', '')} | 国家: {r.get('country', '')} | 产品: {r.get('product', '')}")
        if g and slots:
            lines.append(f"    槽位: intent={slots.get('intent_slot')} | product={slots.get('product_slot')} | scenario={slots.get('scenario_slot')}")
            lines.append(f"          timeline={slots.get('timeline_slot')} | identity={slots.get('identity_strength')} | qty={slots.get('quantity_max')}")
            lines.append(f"    信号: {active_sig}")
            lines.append(f"    L2_TAG={l2_tag} | PriceOnly={price_only}")
            lines.append(f"    Conclusion: {conclusion}")
        else:
            lines.append(f"    ⚠️ 分级失败")
        lines.append(f"    原文: {r.get('content_preview', '')}...")

    # ── 准确率统计 ──
    if labels:
        lines.append(f"\n## 准确率统计")
        lines.append("-" * 40)

        matched = 0
        total_labeled = 0
        confusion = {}

        for r in results:
            g = r.get("grading")
            if not g:
                continue
            email_key = (r.get("email") or "").lower().strip()
            contact_key = (r.get("contact") or "").strip()
            ml = labels.get(email_key) or labels.get(contact_key, "")
            if not ml:
                continue
            total_labeled += 1
            ll = g.get("level", "?")
            if ll == ml:
                matched += 1
            key = f"{ml}→{ll}"
            confusion[key] = confusion.get(key, 0) + 1

        if total_labeled > 0:
            accuracy = matched / total_labeled * 100
            lines.append(f"  有标签样本: {total_labeled}")
            lines.append(f"  整体准确率: {accuracy:.1f}% ({matched}/{total_labeled})")

            for target in ["Level 1", "Level 2"]:
                target_total = sum(1 for r in results
                    if labels.get((r.get("email") or "").lower().strip()) == target
                    or labels.get((r.get("contact") or "").strip()) == target)
                if target_total == 0:
                    continue
                target_hit = sum(1 for r in results
                    if (labels.get((r.get("email") or "").lower().strip()) == target
                        or labels.get((r.get("contact") or "").strip()) == target)
                    and r.get("grading", {}).get("level") == target)
                recall = target_hit / target_total * 100
                lines.append(f"  {target} 召回率: {recall:.1f}% ({target_hit}/{target_total})")

            lines.append(f"\n  混淆矩阵 (人工→模型):")
            for key, count in sorted(confusion.items()):
                lines.append(f"    {key}: {count}")

    return "\n".join(lines)


# ── JSON 输出 ────────────────────────────────────────────────────────────────

def save_results_json(results: list[dict], filepath: str):
    """保存分级结果为 JSON 文件。"""
    output = []
    for r in results:
        entry = {
            "index": r["index"],
            "contact": r.get("contact", ""),
            "email": r.get("email", ""),
            "channel": r.get("channel", ""),
            "country": r.get("country", ""),
            "elapsed": r.get("elapsed", 0),
        }
        if r.get("grading"):
            entry["grading"] = r["grading"]
        output.append(entry)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log.info(f"结果已保存到: {filepath}")


# ── 主入口 ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="线索分级质量验证工具")
    parser.add_argument("--source", help="Excel 或 JSONL 文件路径")
    parser.add_argument("--channel-only", choices=["google", "all"], default="google",
                        help="只处理指定渠道 (default: google)")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 条")
    parser.add_argument("--labels", help="人工标签文件 (JSON/CSV)")
    parser.add_argument("--output", "-o", help="输出 JSON 文件路径")
    parser.add_argument("--delay", type=float, default=0.5, help="LLM 调用间隔秒数 (default: 0.5)")
    parser.add_argument("--stdin", action="store_true", help="从 stdin 读取单条询盘")
    args = parser.parse_args()

    # ── 单条 stdin 模式 ──
    if args.stdin:
        if not os.environ.get("ZHIPU_API_KEY"):
            print("错误: 需要设置 ZIPHU_API_KEY", file=sys.stderr)
            sys.exit(1)
        body = sys.stdin.read().strip()
        if not body:
            print("错误: stdin 为空", file=sys.stderr)
            sys.exit(1)
        result = grade_lead(body)
        if result:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print("分级失败", file=sys.stderr)
            sys.exit(1)
        return

    # ── 批量模式 ──
    if not args.source:
        print("错误: 需要指定 --source 文件", file=sys.stderr)
        sys.exit(1)

    if not os.environ.get("ZHIPU_API_KEY"):
        print("错误: 需要设置 ZHIPU_API_KEY 环境变量", file=sys.stderr)
        sys.exit(1)

    # 加载数据
    if args.source.endswith(".xlsx"):
        leads = load_from_excel(args.source, args.channel_only)
    elif args.source.endswith(".jsonl"):
        leads = load_from_jsonl(args.source)
    else:
        log.error(f"不支持的文件格式: {args.source}")
        sys.exit(1)

    if not leads:
        log.error("没有找到符合条件的询盘")
        sys.exit(1)

    if args.limit > 0:
        leads = leads[:args.limit]
        log.info(f"限制为前 {args.limit} 条")

    labels = None
    if args.labels:
        labels = load_manual_labels(args.labels)
        log.info(f"加载 {len(labels)} 条人工标签")

    # 执行分级
    log.info(f"开始分级 {len(leads)} 条询盘...")
    results = grade_leads(leads, delay=args.delay)

    # 生成报告
    report = generate_report(results, labels)
    print(report)

    # 保存 JSON 结果
    output_path = args.output or f"grading_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    save_results_json(results, output_path)

    log.info("验证完成")


if __name__ == "__main__":
    main()
